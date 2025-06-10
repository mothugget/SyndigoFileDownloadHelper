from openpyxl import load_workbook
wb = load_workbook(filename= "605cf660-bef0-4982-8737-ca068f32faf1_out.xlsm")
ws =wb.active
print(ws['a4'].value)