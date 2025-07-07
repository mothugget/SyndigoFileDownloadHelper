import os
import time
import platform
from pathlib import Path
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook
import zipfile
import xml.etree.ElementTree as ET
import shutil
import time
from lxml import etree as ET
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    print("Warning: python-dotenv not installed. Install with: pip install python-dotenv")


class DownloadHandler(FileSystemEventHandler):
    def __init__(self):
        super().__init__()
        self.recently_processed = set()  # Track recently processed files to avoid loops
        
    def on_created(self, event):
        print(f"🔔 File system event detected: {event.src_path}")
        print(f"   Event type: {'Directory' if event.is_directory else 'File'}")
        print(f"   Path exists: {os.path.exists(event.src_path)}")
        
        # Skip processing if this is a file we just processed (prevents infinite loops)
        if not event.is_directory:
            filename = Path(event.src_path).name
            if filename in self.recently_processed:
                print(f"   ⏭️  Skipping recently processed file: {filename}")
                return
        
        # Skip if file doesn't exist (common on macOS due to rapid file operations)
        if not event.is_directory and not os.path.exists(event.src_path):
            print(f"   ⚠️  File no longer exists, skipping processing")
            return
            
        new_filename = self.process_file(event)
        if new_filename:
            print(f"📝 File renamed via watchdog: {event.src_path} -> {new_filename}")
        
    def on_modified(self, event):
        print(f"📝 File modified: {event.src_path}")
        
    def on_moved(self, event):
        print(f"📦 File moved: {event.src_path} -> {event.dest_path}")
        # Skip processing if this is a file we just renamed (prevents infinite loops)
        dest_filename = Path(event.dest_path).name
        if dest_filename in self.recently_processed:
            print(f"   ⏭️  Skipping recently processed file: {dest_filename}")
            return
            
        # Process the destination file if it's a complete download
        if not event.is_directory and event.dest_path.endswith('.xlsx'):
            # Create a mock event for the destination file
            class MockEvent:
                def __init__(self, path):
                    self.src_path = path
                    self.is_directory = False
            
            new_filename = self.process_file(MockEvent(event.dest_path))
            if new_filename:
                print(f"📝 File renamed via move event: {event.dest_path} -> {new_filename}")
        
    def process_file(self, event):
        """Process a file and return the new filename if renamed, None otherwise"""
        REMOVE_TENANT_ID=os.getenv('REMOVE_TENANT_ID',False)
        MODEL_CONFIGS = {
            "GOVERNANCE MODEL": {
                "prefix": os.getenv('GOVERNANCE_MODEL_PREFIX', 'gov_'), 
                "filename": os.getenv('GOVERNANCE_MODEL_FILENAME', 'governance_model')
            },
            "TAXONOMY APP MODEL": {
                "prefix": os.getenv('TAXONOMY_APP_MODEL_PREFIX', 'tax_'), 
                "filename": os.getenv('TAXONOMY_APP_MODEL_FILENAME', 'taxonomy_model')
            },
            "WORKFLOW APP MODEL": {
                "prefix": os.getenv('WORKFLOW_APP_MODEL_PREFIX', 'wfm_'), 
                "filename": os.getenv('WORKFLOW_APP_MODEL_FILENAME', 'workflow_model')
            },
            "INSTANCE DATA MODEL": {
                "prefix": os.getenv('INSTANCE_DATA_MODEL_PREFIX', 'ins_'), 
                "filename": os.getenv('INSTANCE_DATA_MODEL_FILENAME', 'instance_model')
            },
            "AUTHORIZATION MODEL": {
                "prefix": os.getenv('AUTHORIZATION_MODEL_PREFIX', 'auth_'), 
                "filename": os.getenv('AUTHORIZATION_MODEL_FILENAME', 'authorization_model')
            },
            "KNOWLEDGE DATA MODEL": {
                "prefix": os.getenv('KNOWLEDGE_DATA_MODEL_PREFIX', 'kbm_'), 
                "filename": os.getenv('KNOWLEDGE_DATA_MODEL_FILENAME', 'knowledge_model')
            },
            "RS EXCEL": {
                "prefix": os.getenv('RS_EXCEL_PREFIX', 'data_'), 
                "filename": os.getenv('RS_EXCEL_FILENAME', 'rs_excel_data')
            },
            "thing": {
                "prefix": os.getenv('THING_MODEL_PREFIX', 'thg_'), 
                "filename": os.getenv('THING_MODEL_FILENAME', 'thing_model')
            },
            "referenceData": {
                "prefix": os.getenv('REFERENCEDATA_PREFIX', 'ref_'), 
                "filename": os.getenv('REFERENCEDATA_FILENAME', 'reference_data')
            },
            "UOMData": {
                "prefix": os.getenv('UOMDATA_PREFIX', 'uom_'), 
                "filename": os.getenv('UOMDATA_FILENAME', 'uom_data')
            },
            "digitalAsset": {
                "prefix": os.getenv('DIGITALASSET_PREFIX', 'dam_'), 
                "filename": os.getenv('DIGITALASSET_FILENAME', 'digital_asset')
            },
        }
        try:
            # Only process files, not directories
            if not event.is_directory:
                file_path = Path(event.src_path)
                
                # Double-check file exists before processing (race condition protection)
                if not file_path.exists():
                    print(f"📄 File no longer exists: {file_path.name}")
                    print("-" * 50)
                    return None
                
                file_extension = file_path.suffix.lower()
                already_has_prefix=has_known_prefix(file_path.name,MODEL_CONFIGS)
                
                # Skip files that contain timestamp suffixes (already processed)
                has_timestamp_suffix = "_" in file_path.stem and file_path.stem.split("_")[-1].isdigit()
                
                # Skip _oldv files created during rename process
                has_oldv_suffix = "_oldv" in file_path.stem
                
                print(f"📄 Processing file: {file_path.name}")
                print(f"   Extension: {file_extension}")
                print(f"   Has prefix: {already_has_prefix}")
                print(f"   Has timestamp suffix: {has_timestamp_suffix}")
                print(f"   Has oldv suffix: {has_oldv_suffix}")
                
                if  file_extension in [".xlsx",".xlsm"] and not already_has_prefix and not has_timestamp_suffix and not has_oldv_suffix:
                    print("   ✅ File matches criteria, processing...")
                    try:
                        wb=load_workbook(file_path)
                    except FileNotFoundError:
                        print(f"   ⚠️  File disappeared during processing: {file_path.name}")
                        print("-" * 50)
                        return None
                    ws=wb['METADATA']
                    if ws['A7'].value == 'TENANT' and REMOVE_TENANT_ID:
                        ws['B7'].value = None
                        wb.save(file_path)
                    template_name=str(ws['b4'].value)
                    domain_name=str(ws['b8'].value)
                    model_name=""
                    prefix=""
                    new_file_path=""
                    base_model=False

                    if template_name in MODEL_CONFIGS:
                        model_name=template_name
                    elif domain_name in MODEL_CONFIGS:
                        model_name=domain_name
                        base_model=True

                    print(f"New file detected: {file_path.name}")

                    if model_name!="":
                        replace_filename = os.getenv('REPLACE_FILENAME', 'false').lower() == 'true'
                        prefix = MODEL_CONFIGS[model_name]["prefix"]
                        global_prefix = os.getenv('GLOBAL_PREFIX', '')
                        postfix = os.getenv('FILENAME_POSTFIX', '')
                        
                        if replace_filename:
                            # Replace entire filename with custom name (with global prefix)
                            custom_filename = MODEL_CONFIGS[model_name]["filename"]
                            new_file_path = file_path.parent / (global_prefix + custom_filename + postfix + file_path.suffix)
                        else:
                            # Add prefix to existing filename (with global prefix)
                            base_name = file_path.stem  # filename without extension
                            new_filename = global_prefix + prefix + base_name + postfix + file_path.suffix
                            new_file_path = file_path.parent / new_filename
                        
                        if new_file_path.exists():
                            # Rename existing file to _oldv1, _oldv2, etc.
                            if not rename_existing_file_to_old_version(new_file_path):
                                print(f"   ⚠️  Could not rename existing file. New file will not be processed.")
                                return None
                        
                        file_path.rename(new_file_path)                        
                        print(f"Template name: {template_name}")
                        if file_extension==".xlsx":
                            print(f"Domain name: {domain_name}")
                            disable_window_protection_in_sheetview(new_file_path)
                        print(f"Global prefix: {global_prefix}")
                        print(f"Model prefix: {prefix}")
                        print(f"Postfix: {postfix}")
                        print(f"New filename: {new_file_path.name}")
                        print(f"Replace mode: {replace_filename}")
                        
                        # Move to processed folder if specified
                        final_file_path = move_to_processed_folder(new_file_path)
                        
                        # Track the processed file to prevent infinite loops
                        self.recently_processed.add(final_file_path.name)
                        
                        # Clean up old entries periodically (keep last 100 processed files)
                        if len(self.recently_processed) > 100:
                            oldest_entries = list(self.recently_processed)[:50]
                            for entry in oldest_entries:
                                self.recently_processed.discard(entry)
                        
                        return str(final_file_path)  # Return final filename
                else:
                    print("   ❌ File doesn't match criteria (wrong extension or has prefix)")

                print("-" * 50)    
                return None  # No file was renamed

        except Exception as e:
            print("some file caused some error")
            print(e)
            return None

def disable_window_protection_in_sheetview(xlsx_path):
    xlsx_path = Path(xlsx_path)
    temp_dir = Path("temp_unzip")
    if temp_dir.exists():
        shutil.rmtree(temp_dir)

    # Step 1: Unzip workbook
    with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
        zip_ref.extractall(temp_dir)

    # Step 2: Modify sheet XML files
    sheet_dir = temp_dir / 'xl' / 'worksheets'
    for fpath in sheet_dir.glob('sheet*.xml'):
        parser = ET.XMLParser(remove_blank_text=False)
        tree = ET.parse(str(fpath), parser)
        root = tree.getroot()

        ns = {
            'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
        }
        sheet_view = root.find('.//main:sheetView', namespaces=ns)
        if sheet_view is not None and 'windowProtection' in sheet_view.attrib:
            del sheet_view.attrib['windowProtection']
            print(f"✅ Removed windowProtection in {fpath.name}")
            tree.write(str(fpath), encoding='utf-8', xml_declaration=True, pretty_print=False)

    # Step 3: Repackage - replace original file
    temp_new_file = xlsx_path.with_name(f"{xlsx_path.stem}_temp.xlsx")
    with zipfile.ZipFile(temp_new_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        for file_path in sorted(temp_dir.rglob("*")):
            if file_path.is_file():
                arcname = file_path.relative_to(temp_dir)
                zf.write(file_path, arcname)

    # Step 4: Replace original file
    xlsx_path.unlink()  # Remove original file
    temp_new_file.rename(xlsx_path)  # Rename temp file to original name
    
    # Step 5: Cleanup
    shutil.rmtree(temp_dir)
    print(f"🎉 Window protection removed from: {xlsx_path.name}")



def has_known_prefix(file_name, model_configs) -> bool:
    """
    Returns True if the file name contains any prefix from the model_configs dictionary
    or matches any of the custom filenames.
    
    Parameters:
    - file_name (str): The name of the file (e.g., file_path.name).
    - model_configs (dict): Dictionary with model names and their prefix settings.
    
    Returns:
    - bool: True if any prefix or custom filename is found in the file name, False otherwise.
    """
    file_name_string = str(file_name)
    
    # Check for prefixes
    if any(cfg["prefix"] in file_name_string for cfg in model_configs.values()):
        return True
    
    # Check for custom filenames (when REPLACE_FILENAME=true)
    postfix = os.getenv('FILENAME_POSTFIX', '')
    for cfg in model_configs.values():
        custom_filename = cfg["filename"]
        # Remove file extension for comparison
        name_without_ext = Path(file_name_string).stem
        # Check both with and without postfix
        if name_without_ext == custom_filename or name_without_ext == (custom_filename + postfix):
            return True
    
    return False

def get_downloads_directory():
    """Get the user's downloads directory"""
    # First check .env file
    env_downloads_dir = os.getenv('DOWNLOADS_DIR')
    if env_downloads_dir:
        return env_downloads_dir
    
    # Common download directory paths
    downloads_paths = [
        Path.home() / "Downloads",
        Path.home() / "Download",
        Path.home() / "downloads"
    ]
    
    for path in downloads_paths:
        if path.exists():
            return str(path)
    
    # Fallback: ask user to specify
    custom_path = input("Please enter your downloads directory path: ")
    return custom_path

def add_suffix_to_filename(path: Path, suffix: str = "1") -> Path:
    return path.with_name(f"{path.stem}{suffix}{path.suffix}")

def is_file_locked(file_path: Path) -> bool:
    """Check if a file is locked (open in another application)"""
    # Skip file locking check on macOS due to compatibility issues
    if platform.system() == 'Darwin':
        return False
    
    try:
        # Try to open the file in write mode
        with open(file_path, 'r+b'):
            pass
        return False
    except (IOError, OSError, PermissionError):
        return True

def rename_existing_file_to_old_version(file_path: Path):
    """Rename existing file to next available _oldv# to keep newest file with original name"""
    base_path = file_path.with_suffix('')  # Remove extension
    extension = file_path.suffix
    
    # Check if file is locked (open in Excel/other app)
    if is_file_locked(file_path):
        print(f"🔒 File {file_path.name} is currently open in another application")
        print(f"   Please save and close the file, then it will be automatically renamed")
        
        # Wait for file to be unlocked
        max_attempts = 60  # Wait up to 5 minutes (60 * 5 seconds)
        attempt = 0
        
        while is_file_locked(file_path) and attempt < max_attempts:
            attempt += 1
            print(f"   ⏳ Waiting for file to be closed... (attempt {attempt}/{max_attempts})")
            time.sleep(5)  # Wait 5 seconds between checks
        
        if is_file_locked(file_path):
            print(f"   ❌ Timeout: File still locked after 5 minutes. Skipping rename.")
            print(f"   💡 Please close {file_path.name} manually and run the program again")
            return False
        else:
            print(f"   ✅ File is now available for renaming")
    
    # Find the next available version number
    version = 1
    while True:
        old_version_path = Path(f"{base_path}_oldv{version}{extension}")
        if not old_version_path.exists():
            break
        version += 1
    
    try:
        # Rename the current file to the next available _oldv#
        old_version_path = Path(f"{base_path}_oldv{version}{extension}")
        file_path.rename(old_version_path)
        print(f"📋 Renamed existing {file_path.name} → {old_version_path.name}")
        return True
    except Exception as e:
        print(f"❌ Error renaming file: {e}")
        return False

def move_to_processed_folder(file_path: Path) -> Path:
    """Move processed file to the specified processed files directory if configured"""
    processed_dir = os.getenv('PROCESSED_FILES_DIR', '').strip()
    
    if not processed_dir:
        print("   📁 No processed files directory configured, keeping in place")
        return file_path
    
    try:
        processed_path = Path(processed_dir)
        
        # Create directory if it doesn't exist
        processed_path.mkdir(parents=True, exist_ok=True)
        
        # Create destination path
        destination = processed_path / file_path.name
        
        # Handle filename conflicts
        if destination.exists():
            if not rename_existing_file_to_old_version(destination):
                print(f"   ⚠️  Could not rename existing file in processed folder. File will remain in downloads.")
                return file_path
        
        # Move the file
        shutil.move(str(file_path), str(destination))
        print(f"📦 Moved processed file to: {destination}")
        
        return destination
        
    except Exception as e:
        print(f"❌ Error moving file to processed folder: {e}")
        print(f"   File remains at: {file_path}")
        return file_path

def poll_directory(downloads_dir):
    """Polling-based file monitoring for WSL compatibility"""
    print(f"🔄 Using polling method for WSL compatibility")
    known_files = set()
    handler = DownloadHandler()
    
    # Initialize with existing files
    try:
        for item in Path(downloads_dir).iterdir():
            if item.is_file():  
                known_files.add(str(item))
        print(f"📊 Initial scan found {len(known_files)} files")
    except Exception as e:
        print(f"Error scanning directory: {e}")
        return
    
    try:
        while True:
            current_files = set()
            try:
                for item in Path(downloads_dir).iterdir():
                    if item.is_file():
                        current_files.add(str(item))
                
                # Check for new files
                new_files = current_files - known_files
                if new_files:
                    print(f"🔍 Found {len(new_files)} new files to process")
                
                files_to_add = set()
                files_to_remove = set()
                
                for new_file in new_files:
                    print(f"🆕 New file detected via polling: {new_file}")
                    # Create a mock event object
                    class MockEvent:
                        def __init__(self, path):
                            self.src_path = path
                            self.is_directory = False
                    
                    # Process the file and get the new filename if renamed
                    new_filename = handler.process_file(MockEvent(new_file))
                    
                    if new_filename:
                        # File was renamed, track both old and new names
                        files_to_remove.add(new_file)
                        files_to_add.add(new_filename)
                        print(f"📝 File renamed: {new_file} -> {new_filename}")
                    else:
                        # File was not processed/renamed, just track it normally
                        files_to_add.add(new_file)
                
                # Update known_files properly after processing
                known_files = current_files
                
                # If we processed any files, wait a bit longer before next poll
                if new_files:
                    print("⏳ Waiting extra time after processing files...")
                    time.sleep(3)
                
            except Exception as e:
                print(f"Error during polling: {e}")
            
            time.sleep(2)  # Poll every 2 seconds
            
    except KeyboardInterrupt:
        print("\n🛑 Stopping polling monitor...")

def main():
    downloads_dir = get_downloads_directory().strip('"\'')
    
    print(f"Testing path: {downloads_dir}")
    print(f"Path exists (os.path.exists): {os.path.exists(downloads_dir)}")
    print(f"Path exists (Path.exists): {Path(downloads_dir).exists()}")
    print(f"Is directory: {os.path.isdir(downloads_dir)}")
    
    if not os.path.exists(downloads_dir):
        print(f"Directory {downloads_dir} does not exist!")
        return
    
    print(f"✅ Successfully found downloads directory: {downloads_dir}")
    
    # Check if we're in WSL - handle Windows compatibility
    is_wsl = False
    is_windows_path = False
    
    try:
        # Try Unix-style OS detection (works on Linux/macOS/WSL)
        uname_info = os.uname().release.lower()
        is_wsl = "microsoft" in uname_info or "wsl" in uname_info
        is_windows_path = downloads_dir.startswith("/mnt/")
    except AttributeError:
        # Windows doesn't have os.uname(), detect Windows paths instead
        is_windows_path = len(downloads_dir) > 1 and downloads_dir[1] == ':'
    
    if is_wsl and is_windows_path:
        print("🐧 WSL + Windows path detected, using polling method...")
        poll_directory(downloads_dir)
    else:
        print("🔍 Using watchdog file monitoring...")
        print("Press Ctrl+C to stop monitoring...")
        
        # Create event handler and observer
        event_handler = DownloadHandler()
        observer = Observer()
        observer.schedule(event_handler, downloads_dir, recursive=False)
        
        # Start monitoring
        observer.start()

        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            print("\nStopping monitor...")
            observer.stop()

        observer.join()
        print("Monitor stopped.")

if __name__ == "__main__":
    main()
